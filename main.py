import env
import pso
import plot
import time
import numpy as np
import openpyxl
import math


def denormalize_position(position, xmin, xmax, ymin, ymax):
    x_real = position[0::2] * (xmax - xmin) + xmin
    y_real = position[1::2] * (ymax - ymin) + ymin
    return np.stack([x_real, y_real], axis=-1).flatten()

def reshape(arr):
    res = []
    points = arr.reshape(-1, 2)
    points = points.tolist()
    for i in range(len(points)):
        res.append([points[i][0], points[i][1]])
    return res

def calculate_path_length(path):
    total_length = 0
    for i in range(1, len(path)):
        x1, y1 = path[i - 1]
        x2, y2 = path[i]
        # 计算两点间的欧几里得距离
        total_length += math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
    return total_length


if __name__ == '__main__':

    environment = env.Env()
    start_point = environment.start.tolist()
    goal = environment.goal.tolist()
    plotter = plot.Plotting(start_point, goal, environment.obs_rectangle)

    start_time = time.time()

    pso_planner = pso.PSOPathPlanner(3, 150, 50, c1=1.4962, c2=1.4962, w=0.7298, wdamp=1.0)
    best_solution_normalized, best_length = pso_planner.optimize()
    best_solution = denormalize_position(best_solution_normalized, 0, 1500, 0, 1200)
    best_solution_path = [start_point] + reshape(best_solution) + [goal]

    end_time = time.time()

    print(best_solution_path)
    print("running time is", end_time - start_time)

    # plotter.plot(best_solution_path, "PSO")

    # save data into an Excel file
    # excel_file_path = 'data.xlsx'
    # try:
    #     # 尝试打开现有的 Excel 文件
    #     workbook = openpyxl.load_workbook(excel_file_path)
    #     sheet = workbook.active
    # except FileNotFoundError:
    #     # 如果文件不存在，则创建一个新的工作簿
    #     workbook = openpyxl.Workbook()
    #     sheet = workbook.active
    #
    # def find_first_empty_row(sheet, column):
    #     row = 1  # 行号从1开始
    #     while sheet.cell(row=row, column=column).value is not None:
    #         row += 1
    #     return row
    #
    # # 找到第一列和第二列的第一个空行
    # first_empty_row_a = find_first_empty_row(sheet, 1)  # 第一列 (a)
    # first_empty_row_b = find_first_empty_row(sheet, 2)  # 第二列 (b)
    #
    # # 依次将列表 a 和 b 的值写入空行开始的位置
    #
    # sheet.cell(row=first_empty_row_a, column=1, value=best_length)  # 写入第一列
    # sheet.cell(row=first_empty_row_b, column=2, value=end_time - start_time)  # 写入第二列
    #
    # # 保存 Excel 文件
    # workbook.save(excel_file_path)
    #
    # print(f"数据已保存到 {excel_file_path}")

    print(calculate_path_length(best_solution_path))



